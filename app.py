from flask import Flask, request, render_template, send_from_directory, jsonify
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from waitress import serve
import os
import webbrowser
import edge_tts
import asyncio
import asgiref
import logging

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process():
    if 'file' not in request.files:
        return 'No file part'

    file = request.files['file']
    if file.filename == '':
        return 'No selected file'

    filename = secure_filename(file.filename)
    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))

    excel_file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    try:
        workbook = load_workbook(filename=excel_file_path)
        sheet = workbook.active

        words = [{"en": row[0], "cn": row[1]} for row in sheet.iter_rows(values_only=True) if row[0]]

        return jsonify(words)

    except FileNotFoundError:
        return f"The file '{filename}' was not found in the directory {app.config['UPLOAD_FOLDER']}."
    except Exception as e:
        return f"An error occurred: {e}"

@app.route('/audio/<text>', methods=['GET'])
async def audio(text):
    VOICE = "en-GB-LibbyNeural"
    tts = edge_tts.Communicate(text, VOICE)
    audio_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{text}.mp3")
    await tts.save(audio_path)
    return send_from_directory(app.config['UPLOAD_FOLDER'], f"{text}.mp3")

class IgnoreAudioPathFilter(logging.Filter):
    def filter(self, record):
        return "/audio/" not in record.getMessage()

log = logging.getLogger('werkzeug')
log.addFilter(IgnoreAudioPathFilter())
# 创建一个文件处理器，写到你指定的文件中
file_handler = logging.FileHandler('application.log')

# 创建一个格式器并将其添加到处理器
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)

# 将处理器添加到日志记录器
log.addHandler(file_handler)

# 设置日志记录级别
log.setLevel(logging.INFO)

if __name__ == '__main__':
    serve(app, host='0.0.0.0', port=8080)